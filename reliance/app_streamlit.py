# app_streamlit.py — Merge many inputs → one output (minimal UI, with loading line)
import os, sys, re, traceback, time
from pathlib import Path
from datetime import datetime
import pandas as pd
import streamlit as st

# --- imports / path setup ---
if "." not in sys.path:
    sys.path.insert(0, ".")
if "/mnt/data" not in sys.path:
    sys.path.append("/mnt/data")

import rl_helper as H
from rl_offline_runner import run_offline   # your hardened runner

# ---------- Safe wrapper so missing columns won't crash ----------
_orig_update = H.update_special_remarks_with_article_code
def _safe_update(df):
    for col, default in [("Checking_set", 0), ("SpecialRemarks", ""), ("Article code", "")]:
        if col not in df.columns:
            df[col] = default
    df["Checking_set"]   = pd.to_numeric(df["Checking_set"], errors="coerce").fillna(0).astype(int)
    df["SpecialRemarks"] = df["SpecialRemarks"].astype(str).fillna("")
    df["Article code"]   = df["Article code"].astype(str).fillna("")
    return _orig_update(df)
H.update_special_remarks_with_article_code = _safe_update
# ---------------------------------------------------------------

st.set_page_config(page_title="Reliance — Merge & Export", page_icon="📦", layout="wide")

# ---------- optional CSS (keeps your theme, only layout polish if styles.css exists) ----------
def inject_css(path: str = "styles.css"):
    p = Path(path)
    if p.exists():
        st.markdown(f"<style>{p.read_text(encoding='utf-8')}</style>", unsafe_allow_html=True)
inject_css()

# ---------- tiny UI helpers (block/card layout without hero) ----------
def begin_block(title: str | None = None):
    st.markdown('<div class="block block-gap">', unsafe_allow_html=True)
    if title:
        st.markdown(f'<div class="section-title">{title}</div>', unsafe_allow_html=True)

def end_block():
    st.markdown('</div>', unsafe_allow_html=True)

# ======================== UI ========================
st.title("📦 Reliance — Merge & Export")

begin_block("Upload files")
excel_files = st.file_uploader(
    "Main Order Excel file(s)", type=["xlsx"], accept_multiple_files=True
)
out_prefix = st.text_input(
    "Output name (optional). Leave blank to auto-name.",
    value=""
)
run_btn = st.button("Merge & Export", type="primary", disabled=(not excel_files))
end_block()

# -------- we only create Status/results when needed (no blank blocks) --------
# helpers for reading/merging
def _tmp_dir():
    d = os.path.join(os.getcwd(), "tmp")
    os.makedirs(d, exist_ok=True)
    return d

def _save_tmp(uploaded, name_hint):
    path = os.path.join(_tmp_dir(), name_hint)
    with open(path, "wb") as f:
        f.write(uploaded.getbuffer())
    return path

def _canon(s: str) -> str:
    s = "" if s is None else str(s)
    s = s.replace("\n", " ")
    s = " ".join(s.split())
    return s.strip()

def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    rename = {}
    for c in list(df.columns):
        key = _canon(c)
        key_l = key.lower().replace(" ", "").replace("-", "_")
        target = None
        if key_l in {"itemid","item_id"}: target = "Item Id"
        elif key_l in {"extitemid","ext_item_id","extitemid."}: target = "Ext Item Id"
        elif key_l in {"skuno","sku","skunumber","sku_no","sku.number"}: target = "SKUNo"
        elif key_l in {"articlecode","article_code"}: target = "Article code"
        elif key_l in {"subproductcode","sub_product_code"}: target = "Sub Product Code"
        elif key_l in {"workorderid","work_order_id"}: target = "Work Order Id"
        elif key_l in {"wosrl","wo_srl","wo.serial","wo_srno"}: target = "WO Srl"
        elif key_l in {"qty.1","qty1","qty_1"}: target = "Qty.1"
        elif key_l in {"itemidstone","item_id_stone"}: target = "Item Id Stone"
        elif key_l in {"code"}: target = "Code"
        elif key_l in {"quality"}: target = "QUALITY"
        elif key_l in {"kt","karat"}: target = "KT"
        if target is None: target = key
        rename[c] = target
    out = df.rename(columns=rename)
    all_nan = [c for c in out.columns if out[c].isna().all()]
    if all_nan: out = out.drop(columns=all_nan)
    return out

_HEADER_KEYS = {
    "item id","itemid","item_id","work order id","wo srl","article code","sub product code",
    "ext item id","skuno","sku","trans date","qty","net qty","pure qty","diamond pieces","dia wt","quality","kt","code"
}

def _find_header_row(df_head: pd.DataFrame) -> int | None:
    best_idx, best_hits = None, 0
    for i in range(min(len(df_head), 25)):
        row = df_head.iloc[i].astype(str).tolist()
        hits = sum(1 for cell in row if (cell or "").lower().strip().replace("  "," ") in _HEADER_KEYS)
        if hits > best_hits:
            best_hits, best_idx = hits, i
    return best_idx if best_hits >= 3 else None

def _read_mainorder_excel_autodetect_app(path: str) -> pd.DataFrame:
    xls = pd.ExcelFile(path, engine="openpyxl")
    chosen = None
    for sheet in xls.sheet_names:
        head = pd.read_excel(xls, sheet_name=sheet, header=None, nrows=25, dtype=str)
        hdr_idx = _find_header_row(head)
        if hdr_idx is not None:
            df = pd.read_excel(xls, sheet_name=sheet, header=hdr_idx, dtype=str)
            df = _normalize_columns(df)
            if any(k in df.columns for k in
                   ["Item Id","Ext Item Id","Work Order Id","Article code","Sub Product Code","SKUNo","Code","QUALITY","KT"]):
                chosen = df
                break
    if chosen is None:
        df = pd.read_excel(xls, sheet_name=0, dtype=str)
        chosen = _normalize_columns(df)
    return chosen.dropna(how="all").reset_index(drop=True)

def _drop_unknown_sheets(path: str):
    """Remove any sheet whose name contains 'unknown' (case-insensitive, spaces ignored)."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path)
        norm = lambda s: re.sub(r"\s+", "", (s or "")).lower()
        targets = [n for n in wb.sheetnames if "unknown" in norm(n)]
        if targets and (len(wb.sheetnames) - len(targets)) >= 1:
            for n in targets:
                wb.remove(wb[n])
            wb.save(path)
        return targets
    except Exception as e:
        return [f"<cleanup_failed: {e}>"]

# ======================== RUN ========================
if run_btn and excel_files:
    # Status block (created only when running)
    begin_block("Status")
    status_line = st.empty()
    prog = st.progress(0)
    end_block()

    try:
        # step 1: saving uploads
        status_line.info("Saving uploads…")
        prog.progress(10)
        tmp_paths, names = [], []
        for uploaded in excel_files:
            p = _save_tmp(uploaded, uploaded.name)
            tmp_paths.append(p); names.append(uploaded.name)

        # step 2: reading
        status_line.info("Reading files…")
        prog.progress(35)
        frames = []
        for p in tmp_paths:
            df = _read_mainorder_excel_autodetect_app(p)
            frames.append(df)

        if not frames:
            raise ValueError("No readable rows found in the uploaded files.")

        # step 3: merging
        status_line.info("Merging rows…")
        prog.progress(55)
        merged = pd.concat(frames, ignore_index=True, sort=False)

        # step 4: writing merged input
        status_line.info("Preparing export…")
        prog.progress(70)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        merged_in_path = os.path.join(_tmp_dir(), f"merged_input_{stamp}.xlsx")
        merged.to_excel(merged_in_path, index=False)

        # step 5: run conversion
        status_line.info("Running converter…")
        prog.progress(90)
        prefix = out_prefix.strip() or None
        out_path = run_offline(
            merged_in_path,
            client_name="Reliance",
            output_prefix=prefix,
            style_master_csv=None,
            validator_csv=None,
        )

        # step 6: cleanup
        status_line.info("Finalizing…")
        removed = _drop_unknown_sheets(out_path)
        prog.progress(100)
        status_line.success("Done.")

        # results (shown only when ready; no blank block above)
        begin_block("Result")
        st.write("Files merged:", ", ".join(names))
        st.write("Merged input rows:", len(merged))
        if removed:
            st.write("Removed sheets:", removed)
        st.write("Output:", os.path.basename(out_path))
        with open(out_path, "rb") as f:
            st.download_button("Download XLSX", f, file_name=os.path.basename(out_path))
        end_block()

    except Exception as e:
        status_line.error("Operation failed.")
        begin_block("Error")
        st.error(str(e))
        with st.expander("Show details"):
            st.code(traceback.format_exc())
        end_block()
