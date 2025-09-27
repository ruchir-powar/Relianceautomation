import sys, types, os, re
import pandas as pd

import rl_mapping as _rl_mapping
Clients = types.ModuleType('Clients')
reliance = types.ModuleType('Clients.reliance')
reliance.rl_mapping = _rl_mapping
Clients.reliance = reliance
sys.modules['Clients'] = Clients
sys.modules['Clients.reliance'] = reliance
sys.modules['Clients.reliance.rl_mapping'] = _rl_mapping

import rl_helper as H
import rl_mapping as M
import rl_excelconverter as XL

PRESERVE_ROWS = True


def _canon(s: str) -> str:
    s = '' if s is None else str(s)
    s = s.replace('\n', ' ')
    s = ' '.join(s.split())
    return s.strip()


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    rename = {}
    for c in list(df.columns):
        key = _canon(c)
        key_l = key.lower().replace(' ', '').replace('-', '_')
        target = None
        if key_l in {'itemid', 'item_id'}: target = 'Item Id'
        elif key_l in {'extitemid', 'ext_item_id', 'extitemid.'}: target = 'Ext Item Id'
        elif key_l in {'skuno', 'sku', 'skunumber', 'sku_no', 'sku.number'}: target = 'SKUNo'
        elif key_l in {'articlecode', 'article_code'}: target = 'Article code'
        elif key_l in {'subproductcode', 'sub_product_code'}: target = 'Sub Product Code'
        elif key_l in {'workorderid', 'work_order_id'}: target = 'Work Order Id'
        elif key_l in {'wosrl', 'wo_srl', 'wo.serial', 'wo_srno'}: target = 'WO Srl'
        elif key_l in {'qty.1', 'qty1', 'qty_1'}: target = 'Qty.1'
        elif key_l in {'itemidstone', 'item_id_stone'}: target = 'Item Id Stone'
        elif key_l in {'code'}: target = 'Code'
        elif key_l in {'quality'}: target = 'QUALITY'
        elif key_l in {'kt', 'karat'}: target = 'KT'
        if target is None: target = key
        rename[c] = target
    out = df.rename(columns=rename)
    all_nan = [c for c in out.columns if out[c].isna().all()]
    if all_nan: out = out.drop(columns=all_nan)
    return out


_HEADER_KEYS = {
    'item id', 'itemid', 'item_id', 'work order id', 'wo srl',
    'article code', 'sub product code', 'ext item id', 'skuno',
    'sku', 'trans date', 'qty', 'net qty', 'pure qty',
    'diamond pieces', 'dia wt', 'quality', 'kt', 'code'
}


def _find_header_row(df_head: pd.DataFrame) -> int | None:
    best_idx, best_hits = None, 0
    for i in range(min(len(df_head), 25)):
        row = df_head.iloc[i].astype(str).tolist()
        hits = sum(1 for cell in row if (cell or '').lower().strip().replace('  ', ' ') in _HEADER_KEYS)
        if hits > best_hits:
            best_hits, best_idx = hits, i
    return best_idx if best_hits >= 3 else None


def _read_mainorder_excel_autodetect(path: str) -> pd.DataFrame:
    xls = pd.ExcelFile(path, engine='openpyxl')
    chosen = None
    for sheet in xls.sheet_names:
        head = pd.read_excel(xls, sheet_name=sheet, header=None, nrows=25, dtype=str)
        hdr_idx = _find_header_row(head)
        if hdr_idx is not None:
            df = pd.read_excel(xls, sheet_name=sheet, header=hdr_idx, dtype=str)
            df = _normalize_columns(df)
            if any(k in df.columns for k in [
                'Item Id', 'Ext Item Id', 'Work Order Id',
                'Article code', 'Sub Product Code', 'SKUNo',
                'Code', 'QUALITY', 'KT'
            ]):
                chosen = df
                break
    if chosen is None:
        df = pd.read_excel(xls, sheet_name=0, dtype=str)
        chosen = _normalize_columns(df)
    chosen = chosen.dropna(how='all').reset_index(drop=True)
    return chosen


def _ensure_column(df: pd.DataFrame, name: str, default='') -> pd.DataFrame:
    if name not in df.columns: df[name] = default
    return df


def _derive_kt_from_text(txt: str) -> str:
    if not txt: return ''
    t = txt.upper().replace(' ', '')
    if '18KT' in t or 'GA18' in t or '18K' in t: return '18KT'
    if '14KT' in t or 'GA14' in t or '14K' in t: return '14KT'
    if '9KT' in t or 'GA09' in t or '9K' in t: return '9KT'
    if 'PT95' in t or 'PT' in t: return 'PT95'
    if 'S925' in t or 'S999' in t: return 'SILVER'
    return ''


def mirror_special_remarks(df: pd.DataFrame) -> pd.DataFrame:
    if 'SpecialRemarks' not in df.columns and 'Special Remarks' in df.columns:
        df['SpecialRemarks'] = df['Special Remarks'].astype(str).fillna('')
    elif 'Special Remarks' not in df.columns and 'SpecialRemarks' in df.columns:
        df['Special Remarks'] = df['SpecialRemarks'].astype(str).fillna('')
    elif 'SpecialRemarks' in df.columns and 'Special Remarks' in df.columns:
        sr = df['SpecialRemarks'].astype(str).fillna('')
        srs = df['Special Remarks'].astype(str).fillna('')
        df['SpecialRemarks'] = sr.where(sr.str.len() > 0, srs)
        df['Special Remarks'] = srs.where(srs.str.len() > 0, sr)
    else:
        df['SpecialRemarks'] = ''
        df['Special Remarks'] = ''
    return df


def run_offline(input_xlsx: str, client_name: str = 'Reliance',
                output_prefix: str = None,
                style_master_csv: str = None,
                validator_csv: str = None) -> str:
    if not os.path.exists(input_xlsx):
        raise FileNotFoundError(f"Input Excel not found: {input_xlsx}")

    rl_df = _read_mainorder_excel_autodetect(input_xlsx)

    if PRESERVE_ROWS:
        main = _normalize_columns(rl_df.copy())
    else:
        rl_cleaned = H.helper_reliance(rl_df)
        if not isinstance(rl_cleaned, pd.DataFrame):
            raise TypeError("helper_reliance() did not return a DataFrame.")
        main = _normalize_columns(rl_cleaned)

    if 'Item Id' not in main.columns and 'Ext Item Id' in main.columns:
        main['Item Id'] = main['Ext Item Id']

    if 'QUALITY' not in main.columns and 'Code' in main.columns:
        main['QUALITY'] = main['Code']
    main['StoneQuality'] = main.get('StoneQuality', main.get('QUALITY', '')).fillna('')

    itemid_ser = main.get('Item Id', '').astype(str)
    first_char = itemid_ser.str[0]
    metal_map = first_char.map(M.mapping_for_quality)
    main['Metal'] = metal_map.fillna('')

    main['KT_std'] = main.get('KT', '')
    main['KT_from_metal'] = main['Metal'].apply(_derive_kt_from_text)
    main['KT_from_itemid'] = itemid_ser.apply(_derive_kt_from_text)
    main['KT_final'] = main['KT_std']
    main.loc[main['KT_final'].eq('') & main['KT_from_metal'].ne(''), 'KT_final'] = main['KT_from_metal']
    main.loc[main['KT_final'].eq('') & main['KT_from_itemid'].ne(''), 'KT_final'] = main['KT_from_itemid']

    def _norm_kt(v):
        v = '' if v is None else str(v).upper().strip().replace(' ', '')
        import re
        if re.match(r'^(14|18|9)KT$', v): return v
        m = re.match(r'^(14|18|9)K$', v)
        if m: return m.group(1) + 'KT'
        m = re.match(r'^(14|18|9)$', v)
        if m: return m.group(1) + 'KT'
        if v in {'GA18', 'GA14', 'GA09'}: return v.replace('GA', '') + 'KT'
        if v.startswith('GAWHI18'): return '18KT'
        if v.startswith('GAWHI14'): return '14KT'
        if 'PT95' in v: return 'PT95'
        if v in {'S925', 'S999'}: return 'SILVER'
        return v or ''

    main['KT_final'] = main['KT_final'].apply(_norm_kt)

    main['Tone'] = itemid_ser.str[15].map(H.map_tone) if len(itemid_ser) else ''
    main['CustomerProductionInstruction'] = itemid_ser.map(H.generate_customer_productinstruction)
    main['ItemSize'] = main.apply(H.map_order_group, axis=1)

    if 'Work Order Id' in main.columns:
        main['JOBWORKNUMBER'] = main['Work Order Id']
    elif 'Work Order Id ' in main.columns:
        main['JOBWORKNUMBER'] = main['Work Order Id ']
    else:
        main['JOBWORKNUMBER'] = ''

    def _fmt_kq(kt, q):
        kt = (kt or '').strip(); q = (q or '').strip()
        return f"{kt} {q}".strip()

    main['KT_QUALITY'] = [_fmt_kq(k, q) for k, q in zip(main['KT_final'], main['StoneQuality'])]

    # ---- Safeguards & remarks sync ----
    main = H.stamping_instruct(main)
    main = _ensure_column(main, 'Checking_set', 0)
    main = _ensure_column(main, 'SpecialRemarks', '')
    main = _ensure_column(main, 'Article code', '')
    main = H.update_special_remarks_with_article_code(main)
    main = mirror_special_remarks(main)

    # Bridge: let rename map catch "Special Remarks" -> ItemRefNo
    if 'Special Remarks' not in main.columns and 'SpecialRemarks' in main.columns:
        main['Special Remarks'] = main['SpecialRemarks']

    main.rename(columns=M.RELIANCE_COLUMN_RENAME_MAP, inplace=True)
    main = H.adjust_production_delivery_date(main)

    base, _ = os.path.splitext(input_xlsx)
    output_prefix = output_prefix or f"{base}_processed"
    output_file = XL.process_and_export(main, output_prefix=output_prefix, set_processed=None)

    # Final cleanup: drop any '*unknown*' sheets
    try:
        from openpyxl import load_workbook
        import re as _re
        wb = load_workbook(output_file)
        targets = [n for n in wb.sheetnames if 'unknown' in _re.sub(r'\s+', '', n.lower())]
        if targets and (len(wb.sheetnames) - len(targets)) >= 1:
            for n in targets: wb.remove(wb[n])
            wb.save(output_file)
    except Exception:
        pass

    return output_file
